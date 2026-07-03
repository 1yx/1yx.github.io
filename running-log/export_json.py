#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_json.py — 把训练计划导出为本博客使用的 JSON。

数据来源：同目录下的 2026-2nd-half.xlsx（手动编辑 Excel 后重新生成即可同步网页）。
产物：
  - ../src/data/training/plan.json            训练计划（构建时打包进页面）
  - ../public/running/calendar/activities.json Strava 实际完成（运行时 fetch；仅配置凭据时生成）

运行（uv 自行加载 openpyxl）：
  uv run --with openpyxl python3 running-log/export_json.py

未配置 Strava 凭据（缺少 strava_client.py / .env）时自动跳过 actual 填充，
仅生成 plan.json —— 保证计划数据可独立生成。
"""
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

from generate_plan import parse_km
from strava_feed import build_feed_entry

# 路径都以本脚本所在目录 (running-log/) 为基准，CWD 无关。
ROOT = Path(__file__).resolve().parent          # running-log/
REPO = ROOT.parent                               # 博客仓库根
WORKBOOK = ROOT / "2026-2nd-half.xlsx"
PLAN_OUT = REPO / "src" / "data" / "training" / "plan.json"
ACTIVITIES_OUT = REPO / "public" / "running" / "calendar" / "activities.json"
DCACHE = ROOT / "strava_activity_details.json"   # 活动详情缓存（按 id 增量）

DAYS_PER_CYCLE = 4
ROWS_PER_CYCLE = 3
CATEGORY_COLORS = {
    "Specific": "#F6A9EE",
    "Special": "#C5A6FF",
    "Fundamental": "#74DDFF",
    "Regeneration": "#CFEDE4",
    "Race": "#FFD700",
    "Rest": "#9AA7BA",
}


def cycle_to_mesocycle_set(cycle):
    return (cycle - 1) // 4 + 1, (cycle - 1) % 4 + 1


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_planned_km(text):
    if not text:
        return 0.0
    if "marathon race" in text.lower():
        return 42.2
    return round(parse_km(text), 1)


def workout_category(phase_name, text, day, period):
    if not text:
        return "Rest"
    if "marathon race" in text.lower():
        return "Race"
    if "Recovery" in text or "Shakeout" in text or "Walk / Jog" in text:
        return "Regeneration"
    if phase_name == "Aerobic Phase":
        if period == "AM" and day in (1, 3):
            return "Fundamental"
    if phase_name == "Threshold Phase":
        if day == 1 and period == "AM":
            return "Specific"
        if (day == 1 and period == "PM") or (day == 3 and period == "PM"):
            return "Special"
        if period == "AM" and day in (2, 3):
            return "Fundamental"
    if phase_name == "Marathon Phase":
        if day == 1 and period == "AM":
            return "Specific"
        if period == "AM" and day in (2, 3):
            return "Fundamental"
    return "Regeneration"


def workout_obj(text, phase_name, cycle, day, period):
    """构造单个 workout 对象：文本 + 配色 + 计划里程 + actual 占位(留给 Strava)。"""
    text = normalize_text(text)
    category = workout_category(phase_name, text, day, period)
    return {
        "text": text,
        "category": category,
        "color": CATEGORY_COLORS[category],
        "planned_km": parse_planned_km(text),   # 计划里程(用于「执行」统计)
        "actual": None,   # 由 fill_actual_from_strava 填充（若有 Strava 凭据）
    }


def phase_name(name):
    return name.replace(" Session", " Phase").replace(" Sheet", " Phase")


def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return datetime.fromisoformat(value).date().isoformat()
    raise ValueError(f"无法解析日期: {value!r}")


def row_has_cycle(ws, row):
    value = ws.cell(row=row, column=1).value
    return isinstance(value, str) and value.strip().startswith("小周期")


def build_phase_from_sheet(ws):
    """一个 Excel sheet -> {name, mesocycles:[{index, microcycles:[...]}]}。"""
    display_name = ws.title
    mesocycles = {}
    cycle = 0
    for header_row in range(1, ws.max_row + 1, ROWS_PER_CYCLE):
        if not row_has_cycle(ws, header_row):
            continue
        cycle += 1
        mesocycle_idx, microcycle_in_mesocycle = cycle_to_mesocycle_set(cycle)
        am_row = header_row + 1
        pm_row = header_row + 2
        days = []
        for day in range(1, DAYS_PER_CYCLE + 1):
            col = day + 1
            days.append({
                "day": day,
                "date": iso_date(ws.cell(row=header_row, column=col).value),
                "am": workout_obj(ws.cell(row=am_row, column=col).value, display_name, cycle, day, "AM"),
                "pm": workout_obj(ws.cell(row=pm_row, column=col).value, display_name, cycle, day, "PM"),
            })
        vals = [d["am"]["text"] for d in days] + [d["pm"]["text"] for d in days]
        total_cell = normalize_text(ws.cell(row=header_row, column=6).value)
        total = sum(parse_planned_km(v) for v in vals)
        is_marathon = display_name == "Marathon Phase"
        microcycle_obj = {
            "index": cycle,                       # 绝对 microcycle 序号
            "in_mesocycle": microcycle_in_mesocycle,      # Mesocycle 内第几个 Microcycle (1-4)
            "cutback": microcycle_in_mesocycle == 4,   # 每个 Mesocycle 的第 4 个 Microcycle = 减量
            "taper": is_marathon and cycle >= 16,
            "total_km": total_cell or (f"{total:g}km" if total else None),
            "days": days,
        }
        mesocycles.setdefault(mesocycle_idx, {"index": mesocycle_idx, "microcycles": []})["microcycles"].append(microcycle_obj)
    mesocycle_list = [mesocycles[k] for k in sorted(mesocycles)]
    return {"name": phase_name(display_name), "mesocycles": mesocycle_list}


def build_data_from_workbook(path=WORKBOOK):
    wb = openpyxl.load_workbook(path, data_only=False)
    phases = []
    for sheet_name in ("Aerobic Phase", "Threshold Phase", "Marathon Phase"):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"缺少工作表: {sheet_name}")
        phases.append(build_phase_from_sheet(wb[sheet_name]))
    return {"phases": phases}


def fill_actual_from_strava(data):
    """若有 strava_client.py 且配置凭据(.env)，拉取活动并回填 actual + 写 activities.json。"""
    try:
        import strava_client
    except ModuleNotFoundError:
        print("未找到 strava_client.py，跳过 Strava actual 填充（仅生成计划数据）。")
        return
    from datetime import timedelta

    if not strava_client.configured():
        print("未配置 Strava 凭据(.env)，跳过 actual 填充（仅生成计划数据）。")
        return
    dates = [d["date"] for phase in data["phases"] for b in phase["mesocycles"]
             for microcycle in b["microcycles"] for d in microcycle["days"]]
    if not dates:
        return
    d0, d1 = min(dates), max(dates)
    after = int((datetime.fromisoformat(d0) - timedelta(days=1)).timestamp())
    before = int((datetime.fromisoformat(d1) + timedelta(days=2)).timestamp())
    print(f"拉取 Strava 活动 {d0} ~ {d1} ...")
    acts = strava_client.list_activities(after_epoch=after, before_epoch=before)
    runs = [a for a in acts if a.get("sport_type") in strava_client.RUN_SPORTS]
    # 按 (日期, AM/PM) 聚合
    agg = {}
    for a in runs:
        sdt = (a.get("start_date_local") or "").replace("Z", "")
        try:
            dt = datetime.fromisoformat(sdt)
        except ValueError:
            continue
        key = (dt.date().isoformat(), "AM" if dt.hour < 12 else "PM")
        v = agg.setdefault(key, {"distance_km": 0.0, "moving_time": 0, "elev_m": 0.0, "n": 0, "name": ""})
        v["distance_km"] += (a.get("distance") or 0) / 1000
        v["moving_time"] += a.get("moving_time") or 0
        v["elev_m"] += a.get("total_elevation_gain") or 0
        v["n"] += 1
        v["name"] = a.get("name", v["name"])
    # 回填到对应计划格
    filled = 0
    for phase in data["phases"]:
        for b in phase["mesocycles"]:
            for microcycle in b["microcycles"]:
                for d in microcycle["days"]:
                    for period in ("am", "pm"):
                        w = d[period]
                        if not w.get("text"):
                            continue
                        v = agg.get((d["date"], period.upper()))
                        if v:
                            w["actual"] = {
                                "distance_km": round(v["distance_km"], 2),
                                "moving_time": v["moving_time"],
                                "elev_m": int(round(v["elev_m"])),
                                "name": v["name"],
                            }
                            filled += 1
    print(f"Strava 回填 {filled} 个 workout 的 actual（活动 {len(acts)} 条，跑步 {len(runs)} 条）")

    # description / splits_metric 列表接口不返回，需逐个取详情(GET /activities/{id})；
    # 按 id 增量缓存，避免重复请求。
    details = {}
    try:
        with open(DCACHE, encoding="utf-8") as f:
            details = json.load(f)
    except FileNotFoundError:
        pass

    def detail_needs_refresh(activity):
        activity_id = str(activity.get("id"))
        cached = details.get(activity_id)
        # v:2 = 缓存的是原始 splits（由 build_feed_entry 处理）；旧版存的是已处理结果，需重取
        return (
            activity.get("id") is not None and
            (not isinstance(cached, dict) or cached.get("v") != 2)
        )

    missing = [a for a in runs if detail_needs_refresh(a)]
    for a in missing:
        try:
            det = strava_client.get_activity(a["id"])
            details[str(a["id"])] = {
                "description": (det.get("description") or "").strip(),
                "splits_metric": det.get("splits_metric", []),   # 存原始，由 build_feed_entry 处理
                "v": 2,
            }
        except Exception as e:
            print(f"  活动 {a['id']} 详情失败: {e}")
            details[str(a["id"])] = {"description": "", "splits_metric": [], "v": 2}
    if missing:
        with open(DCACHE, "w", encoding="utf-8") as f:
            json.dump(details, f, ensure_ascii=False, indent=2)

    feed = []
    for a in runs:
        entry = build_feed_entry(a, details.get(str(a.get("id")), {}))
        if entry is not None:
            feed.append(entry)
    feed.sort(key=lambda x: x["date"] + x["time"], reverse=True)
    ACTIVITIES_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(ACTIVITIES_OUT, "w", encoding="utf-8") as f:
        json.dump(feed, f, ensure_ascii=False, separators=(",", ":"))   # 压缩，运行时 fetch
    print(f"已写出 {ACTIVITIES_OUT.relative_to(REPO)}（{len(feed)} 条跑步活动，"
          f"描述缓存 {len(details)} 条，本次新增 {len(missing)} 条）")


def main():
    data = build_data_from_workbook()
    fill_actual_from_strava(data)
    PLAN_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(PLAN_OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    n_microcycles = sum(len(b["microcycles"]) for phase in data["phases"] for b in phase["mesocycles"])
    print(f"已写出 {PLAN_OUT.relative_to(REPO)}（{len(data['phases'])} phases, {n_microcycles} microcycles）")


if __name__ == "__main__":
    main()
