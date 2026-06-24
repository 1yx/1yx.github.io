#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_plan.py — 从零生成「2026 夏训训练计划」Excel 工作簿。

输出三个 Session（工作表）：Aerobic Session / Threshold Session / Marathon Sheet。
本脚本不依赖任何已有文件，可直接生成完整工作簿，便于下个赛季改参数后复用。

运行：uv run --with openpyxl python3 generate_plan.py
输出：2026夏训-训练计划_generated.xlsx

================================================================================
一、整体结构（层级：Session > Block > Set > Day > Workout）
================================================================================
  Session ：一张工作表 = 一种课表（Aerobic / Threshold / Marathon）。
  Block   ：每 4 个小周期(microcycle)为一块。
  Set     ：块内第 1/2/3 个小周期为 Normal，第 4 个为 Deload(减量)。
  Day     ：每个小周期 4 天（Day1-Day4），占 3 行：
              第 1 行 = 表头(小周期标签 + 4 天日期 + F 列总量)
              第 2 行 = 上午(AM)
              第 3 行 = 下午(PM)
  Workout ：某天上午/下午格里的具体训练文本。

  小周期序号 cycle 与 (Block, Set) 的换算：
      block = (cycle-1)//4 + 1 ； set = (cycle-1)%4 + 1
  减量判定：cycle % 4 == 0（即每块的 Set 4）。

================================================================================
二、各 Session 训练内容与关键变化（改参数时主要看这里）
================================================================================
【Aerobic Session】28 周期，2026/6/1 起。里程随周期递增(ramp，单位 km)：
      cycle 1-8 → +0；9-16 → +1；17-23 → +2；24-28 → +3
    Day1 上午：26+Δ Long Run（Normal）/ 18+Δ Medium Long Run（Deload）
    Day2 上/下午：6km Recovery ×2
    Day3 上午：21+Δ Medium-Long Run（Normal）/ 13+Δ Medium Run（Deload）
    Day4 上/下午：6km Recovery ×2
    Day1/Day3 下午：留空（休息）

【Threshold Session】12 周期(3 块)，2026/9/21 起。每块配速递进 −10″：
      块1=4′00″，块2=3′50″，块3=3′40″
    Day1 上午(Normal Set1/2/3)：2km Warm Up + [4*3km / 2*5.5km / 10km]@块配速
    Day1 下午：5km Warm Up + 3*1km@「下一块」配速（块3 已无下一块，外推 3′30″）
    Day2：上午 10km Medium Run，下午 6km Recovery
    Day3：上午 21km Medium Long Run，下午 6km Warm Up + 5*200m Up Hill
    Day4：上午 6km Recovery，下午 留空
    Deload(Set4)：照抄 Aerobic 周期4 减量，并把 Day4 下午的 Recovery 挪到 Day3 下午。

【Marathon Sheet】18 周期(4 块 + 末尾 taper)，2026/11/8 起（到 2027/1/18）。
    Day1 上午(Normal Set1/2/3，三种都恰好 32km)：
        Set1(各块相同)：16km@Aerobic+16km@Progression
        Set2：[27+5 / 26+6 / 25+7 / 24+8]km@(Aerobic+MP)，对应块1-4
        Set3：4*（[4+4 / 3+5 / 2+6 / 1+7]km@(slow+fast)），对应块1-4
    Day2 / Day3：上午 10km Medium Run，下午 6km Recovery（两天相同）
    Day4：上午 7km Recovery，下午 留空
    Day1 下午：留空
    Deload(Set4)：照抄 Threshold Set4（即上面那套减量）。
    Taper(末尾 2 个小周期)：全部留空。

================================================================================
三、统一配色（三张表共用，背景色）
================================================================================
    表头-常规周期      深灰 595959 + 白字
    表头-减量/taper    中灰 808080 + 白字
    Normal Day1 上午   粉   FAE7E5 (250,231,229)   ← 三张表的主课
    Long Run           粉   FAE7E5                  ← 同上
    Medium Long Run    奶油 FCF2E9 (252,242,233)
    Medium Run         浅黄 FEF8E1 (254,248,225)
    Recovery           蓝灰 F2F3F6 (242,243,246)
    Sprint             紫   F9E9FE (249,233,254)   ← Day1下午 & Day3下午
================================================================================
"""

import re
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================================
# 0. 输出路径（可改）
# ============================================================================
OUTPUT = "2026夏训-训练计划_generated.xlsx"


# ============================================================================
# 1. 统一配色（见文档「三、统一配色」）
# ============================================================================
COLOR_HEADER_DARK = "FF595959"   # 表头-常规：深灰
COLOR_HEADER_MED  = "FF808080"   # 表头-减量/taper：中灰
WHITE             = "FFFFFFFF"   # 表头字体：白

COLOR_DAY1AM   = "FFFAE7E5"      # Normal Day1 上午主课 / Long Run：粉
COLOR_LONG     = "FFFAE7E5"      # （与 DAY1AM 同色）
COLOR_MEDLONG  = "FFFCF2E9"      # Medium Long Run：奶油
COLOR_MEDIUM   = "FFFEF8E1"      # Medium Run：浅黄
COLOR_RECOVERY = "FFF2F3F6"      # Recovery：蓝灰
COLOR_SPRINT   = "FFF9E9FE"      # Sprint：紫
NO_FILL        = None            # 无填充 = 白底


# ============================================================================
# 2. 通用结构常量
# ============================================================================
ROWS_PER_CYCLE = 3       # 每个小周期占 3 行（表头 / 上午 / 下午）
DAYS_PER_CYCLE = 4       # 每个小周期 4 天
DATE_FMT = "yyyy/m/d;@"  # 日期显示格式

# 列定义：A=小周期标签, B=Day1, C=Day2, D=Day3, E=Day4, F=总量
COL_LABEL, COL_DAY1, COL_DAY2, COL_DAY3, COL_DAY4, COL_TOTAL = 1, 2, 3, 4, 5, 6
DAY_TO_COL = {1: COL_DAY1, 2: COL_DAY2, 3: COL_DAY3, 4: COL_DAY4}


def header_row(cycle):
    """cycle 的表头行号。"""
    return 3 * cycle - 2

def am_row(cycle):
    """cycle 的上午行号。"""
    return 3 * cycle - 1

def pm_row(cycle):
    """cycle 的下午行号。"""
    return 3 * cycle

def cycle_to_block_set(cycle):
    """小周期序号 -> (Block, Set)。"""
    return (cycle - 1) // 4 + 1, (cycle - 1) % 4 + 1

def is_deload(cycle):
    """Set 4 = 减量周期。"""
    return cycle % 4 == 0


# ============================================================================
# 3. 训练内容（按 Session > Block > Set > Day > Workout）
#    每个 resolver 给定 (cycle, block, set, day, slot) 返回该格文本，None=留空。
# ============================================================================

# ---------- 3.1 Aerobic Session ----------
AEROBIC_START  = date(2026, 6, 1)
AEROBIC_CYCLES = 28

def aerobic_delta(cycle):
    """Aerobic 里程递增量(km)：1-8→0, 9-16→1, 17-23→2, 24-28→3。"""
    if cycle <= 8:  return 0
    if cycle <= 16: return 1
    if cycle <= 23: return 2
    return 3

def aerobic_workout(cycle, day, slot):
    """Aerobic 某(周期, 天, 上午/下午)的训练内容。"""
    delta = aerobic_delta(cycle)
    deload = is_deload(cycle)
    if day == 1:
        if slot == "AM":
            return f"{18 + delta}km Medium Long Run" if deload else f"{26 + delta}km Long Run"
        return None                       # Day1 下午休息
    if day in (2, 4):                     # Day2 / Day4：上下午都是 Recovery
        return "6km Recovery"
    if day == 3:
        if slot == "AM":
            return f"{13 + delta}km Medium Run" if deload else f"{21 + delta}km Medium-Long Run"
        return None                       # Day3 下午休息
    return None


# ---------- 3.2 Threshold Session ----------
THRESHOLD_START  = date(2026, 9, 21)
THRESHOLD_CYCLES = 12                     # 3 块 × 4
THRESHOLD_PACE      = {1: "4′00″", 2: "3′50″", 3: "3′40″"}    # 各块 Day1 上午配速
THRESHOLD_NEXT_PACE = {1: "3′50″", 2: "3′40″", 3: "3′30″"}    # 下一块配速(Day1下午用；块3外推)
THRESHOLD_INTERVAL  = {1: "4*3km", 2: "2*5.5km", 3: "10km"}   # 块内 Set1/2/3 的上午间歇

def deload_workout(day, slot):
    """减量周期(Set4)内容：照抄 Aerobic 周期4，并把 Day4 下午的 Recovery 挪到 Day3 下午。
    Threshold 与 Marathon 的减量周期共用此函数。"""
    if day == 1: return "18km Medium Long Run" if slot == "AM" else None
    if day == 2: return "6km Recovery"                                # 上下午都是
    if day == 3: return "13km Medium Run" if slot == "AM" else "6km Recovery"  # 下午由 Day4 挪来
    if day == 4: return "6km Recovery" if slot == "AM" else None     # 下午已挪走
    return None

def threshold_workout(cycle, block, set_, day, slot):
    """Threshold 某(周期,块,Set,天,上午/下午)的训练内容。"""
    if set_ == 4:                         # Set4 = 减量
        return deload_workout(day, slot)
    if day == 1:
        if slot == "AM":
            return f"2km Warm Up + {THRESHOLD_INTERVAL[set_]}@{THRESHOLD_PACE[block]}"
        return f"5km Warm Up + 3*1km@{THRESHOLD_NEXT_PACE[block]}"
    if day == 2:
        return "10km Medium Run" if slot == "AM" else "6km Recovery"
    if day == 3:
        if slot == "AM":
            return "21km Medium Long Run"
        return "6km Warm Up + 5*200m Up Hill"           # 冲刺课
    if day == 4:
        return "6km Recovery" if slot == "AM" else None
    return None


# ---------- 3.3 Marathon Sheet ----------
MARATHON_START  = date(2026, 11, 8)
MARATHON_CYCLES = 18                     # 4 块 × 4 + 末尾 taper(17,18)
# 各块 Set1/2/3 的 Day1 上午内容（三种都恰好 32km）
MARATHON_DAY1 = {
    (1, 1): "16km@Aerobic+16km@Progression", (1, 2): "27km@Aerobic+5km@MP", (1, 3): "4*（4km@slow+4km@fast）",
    (2, 1): "16km@Aerobic+16km@Progression", (2, 2): "26km@Aerobic+6km@MP", (2, 3): "4*（3km@slow+5km@fast）",
    (3, 1): "16km@Aerobic+16km@Progression", (3, 2): "25km@Aerobic+7km@MP", (3, 3): "4*（2km@slow+6km@fast）",
    (4, 1): "16km@Aerobic+16km@Progression", (4, 2): "24km@Aerobic+8km@MP", (4, 3): "4*（1km@slow+7km@fast）",
}

def marathon_workout(cycle, block, set_, day, slot):
    """Marathon 某(周期,块,Set,天,上午/下午)的训练内容。"""
    if set_ == 4:                         # Set4 = 减量（同 Threshold Set4）
        return deload_workout(day, slot)
    if (block, set_) not in MARATHON_DAY1:    # taper(块5) → 留空
        return None
    if day == 1:
        return MARATHON_DAY1[(block, set_)] if slot == "AM" else None
    if day in (2, 3):
        return "10km Medium Run" if slot == "AM" else "6km Recovery"
    if day == 4:
        return "7km Recovery" if slot == "AM" else None
    return None


# ============================================================================
# 4. 配色规则
# ============================================================================
def header_fill_rgb(session_name, cycle, n_cycles):
    """表头行底色：减量 或 Marathon 末尾 taper 用中灰，其余深灰。"""
    if is_deload(cycle):
        return COLOR_HEADER_MED
    if session_name == "Marathon Sheet" and cycle >= n_cycles - 1:   # 末尾 2 周期 = taper
        return COLOR_HEADER_MED
    return COLOR_HEADER_DARK

def workout_fill_rgb(session_name, cycle, day, slot, text):
    """训练格底色。位置覆盖优先（Day1上午主课/Threshold冲刺课），其次按内容关键词。"""
    _, set_ = cycle_to_block_set(cycle)
    is_normal = set_ in (1, 2, 3)
    # --- 1) 位置覆盖 ---
    if is_normal:
        if day == 1 and slot == "AM":
            return COLOR_DAY1AM                 # Normal Day1 上午主课 = 粉（三张表统一）
        if session_name == "Threshold Session" and \
           ((day == 1 and slot == "PM") or (day == 3 and slot == "PM")):
            return COLOR_SPRINT                  # Threshold 冲刺课 = 紫
    # --- 2) 内容关键词 ---
    if text:
        low = text.lower()
        if "recovery" in low:                   return COLOR_RECOVERY
        if "medium" in low and "long" in low:   return COLOR_MEDLONG
        if "medium" in low:                     return COLOR_MEDIUM
        if "long" in low:                       return COLOR_LONG
    return NO_FILL                              # 其余（空格/阈值课等）= 白


# ============================================================================
# 5. 跑量解析（用于 F 列总量）
#    支持：N*（A+…）、A*Bkm、A*Bm、Akm、Am；乘法片段会覆盖其所含的子匹配，避免重复。
# ============================================================================
_RE_PAREN = re.compile(r"(\d+(?:\.\d+)?)\s*\*\s*[（(]([^）)]+)[）)]")
_RE_MKM   = re.compile(r"(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)\s*km")
_RE_MM    = re.compile(r"(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)\s*m\b")
_RE_KM    = re.compile(r"(\d+(?:\.\d+)?)\s*km")
_RE_M     = re.compile(r"(\d+(?:\.\d+)?)\s*m\b")

def parse_km(text):
    """解析训练文本中的跑量(km)。"""
    if not isinstance(text, str):
        return 0.0
    total = 0.0
    used = []
    for m in _RE_PAREN.finditer(text):                       # N*（内含求和）
        inner = sum(float(x.group(1)) for x in _RE_KM.finditer(m.group(2)))
        total += float(m.group(1)) * inner
        used.append((m.start(), m.end()))
    for m in _RE_MKM.finditer(text):                         # A*Bkm
        if any(s <= m.start() < e for s, e in used):
            continue
        total += float(m.group(1)) * float(m.group(2))
        used.append((m.start(), m.end()))
    for m in _RE_MM.finditer(text):                          # A*Bm
        if any(s <= m.start() < e for s, e in used):
            continue
        total += float(m.group(1)) * float(m.group(2)) / 1000
        used.append((m.start(), m.end()))
    for m in _RE_KM.finditer(text):                          # Akm
        if any(s <= m.start() < e for s, e in used):
            continue
        total += float(m.group(1))
    for m in _RE_M.finditer(text):                           # Am
        if any(s <= m.start() < e for s, e in used):
            continue
        total += float(m.group(1)) / 1000
    return total


# ============================================================================
# 6. 构建工具
# ============================================================================
def solid(rgb):
    """返回指定颜色的纯色填充。"""
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

def set_column_widths(ws):
    """设置列宽（A 标签 / B-E 训练格较宽 / F 总量较窄）。"""
    for col, width in {1: 14, 2: 30, 3: 26, 4: 26, 5: 26, 6: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = width

def set_white_font(cell):
    """把单元格字体改为白色（保留字体名/字号/粗斜体等其它属性）。"""
    f = cell.font
    cell.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                     underline=f.underline, strikethrough=f.strikethrough, color=WHITE)

def right_align_all(ws, n_cycles):
    """表区内(rows 1..3n, cols A-F)所有单元格水平右对齐；顺带补齐空单元格
    （如 A 列/F 列的内容行），使整张表对齐与手动编辑版一致。"""
    for r in range(1, 3 * n_cycles + 1):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            a = cell.alignment
            cell.alignment = Alignment(horizontal="right", vertical=a.vertical,
                                       wrap_text=a.wrap_text, text_rotation=a.text_rotation,
                                       indent=a.indent, shrink_to_fit=a.shrink_to_fit)


# 三个 Session 的配置：name / start / n_cycles / workout 解析函数
# workout 解析函数签名统一为 (cycle, block, set, day, slot) -> 文本或 None
SESSIONS = [
    {"name": "Aerobic Session",   "start": AEROBIC_START,   "n": AEROBIC_CYCLES,
     "workout": lambda c, b, s, d, sl: aerobic_workout(c, d, sl)},
    {"name": "Threshold Session", "start": THRESHOLD_START, "n": THRESHOLD_CYCLES,
     "workout": lambda c, b, s, d, sl: threshold_workout(c, b, s, d, sl)},
    {"name": "Marathon Sheet",    "start": MARATHON_START,  "n": MARATHON_CYCLES,
     "workout": lambda c, b, s, d, sl: marathon_workout(c, b, s, d, sl)},
]


def build_session(wb, session):
    """构建一个 Session 工作表。
    步骤（与手动编辑时一致）：
      ① 表头行：A=小周期标签、B:E=日期、F=总量(后填) → 上灰底白字
      ② 训练内容：B:E 的上午/下午行 → 按规则填文本+底色
      ③ F 列：解析本周期各格跑量求和（无内容如 taper 则留空）
      ④ 整表右对齐
    """
    name = session["name"]
    start = session["start"]
    n = session["n"]
    workout_fn = session["workout"]

    ws = wb.create_sheet(name)
    set_column_widths(ws)

    for cycle in range(1, n + 1):
        block, set_ = cycle_to_block_set(cycle)     # 该周期对应的 Block / Set
        hr, ar, pr = header_row(cycle), am_row(cycle), pm_row(cycle)

        # --- ① 表头行：标签 + 日期 ---
        ws.cell(row=hr, column=COL_LABEL, value=f"小周期{cycle}")
        d0 = start + timedelta(days=(cycle - 1) * DAYS_PER_CYCLE)   # 本周期 Day1 的日期
        for day in range(1, DAYS_PER_CYCLE + 1):
            c = ws.cell(row=hr, column=DAY_TO_COL[day],
                        value=d0 + timedelta(days=day - 1))
            c.number_format = DATE_FMT
        # 表头底色 + 白字（A:F 整行）
        hfill = solid(header_fill_rgb(name, cycle, n))
        for col in range(1, 7):
            ws.cell(row=hr, column=col).fill = hfill
            set_white_font(ws.cell(row=hr, column=col))

        # --- ② 训练内容 + 底色 ---
        for day in range(1, DAYS_PER_CYCLE + 1):
            col = DAY_TO_COL[day]
            for slot, row in (("AM", ar), ("PM", pr)):
                text = workout_fn(cycle, block, set_, day, slot)
                cell = ws.cell(row=row, column=col, value=text)
                rgb = workout_fill_rgb(name, cycle, day, slot, text)
                cell.fill = solid(rgb) if rgb else PatternFill()   # 无颜色=默认白底

        # --- ③ F 列总量 ---
        cells = [ws.cell(row=r, column=c).value
                 for r in (ar, pr) for c in (COL_DAY1, COL_DAY2, COL_DAY3, COL_DAY4)]
        if any(isinstance(v, str) and v.strip() for v in cells):   # 有训练内容才算
            total = sum(parse_km(v) for v in cells)
            ws.cell(row=hr, column=COL_TOTAL, value=f"{total:g}km")
        # 无内容（如 Marathon taper）→ F 留空

    # --- ④ 全部右对齐 ---
    right_align_all(ws, n)
    print(f"[{name}] 完成：{n} 个小周期")


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)                 # 删除默认空表
    for session in SESSIONS:             # 依次构建三个 Session
        build_session(wb, session)
    wb.save(OUTPUT)
    print(f"\n已生成 -> {OUTPUT}")


if __name__ == "__main__":
    main()
